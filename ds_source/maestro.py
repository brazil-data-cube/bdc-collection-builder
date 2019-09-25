import os
import io
from flask import Flask, request, make_response, render_template, abort, jsonify
from redis import Redis, RedisError
import utils
from  utils import c2jyd,do_insert,do_update,do_upsert,do_query,do_command,decodePeriods,decodePathRow
import sqlalchemy
import time
import datetime
import fnmatch
import glob
import logging
import requests
import json
import threading
import numpy
import openpyxl
from openpyxl.chart import ScatterChart, Reference, Series
from openpyxl.comments import Comment
from flask_cors import CORS

app = Flask(__name__)
cors = CORS(app, resources={r"/*": {"origins": "*"}})
app.config['PROPAGATE_EXCEPTIONS'] = True
app.logger_name = "maestro"
handler = logging.FileHandler('maestro.log')
handler.setFormatter(logging.Formatter(
	'[%(asctime)s] %(levelname)s in %(module)s: %(message)s'
))

app.logger.addHandler(handler)
app.jinja_env.trim_blocks = True
app.jinja_env.lstrip_blocks = True
app.jinja_env.keep_trailing_newline = True

redis = Redis(host="redis", db=0)

MAX_THREADS = int(os.environ.get('MAX_THREADS'))
CUR_THREADS = 0
ACTIVITIES = {'search':{'current':0,'maximum':8},'download':{'current':0,'maximum':8},'warp':{'current':0,'maximum':16},'merge':{'current':0,'maximum':16},'blend':{'current':0,'maximum':8}}

###################################################
def getLock():
# Manage lock
	app.logger.warning('getLock from Redis.')
	lock = None
	try:
		lock = redis.get('lock')
	except RedisError:
		app.logger.exception('Cannot connect to Redis.')
		return 0
	if lock is None:
		redis.set('lock',0)

	return redis.get('lock')


###################################################
def sendToSoloist(activity):
# Sending command to soloist
	cmd = ''
	for key,val in activity.items():
		if val is not None:
			cmd += key+'='+str(val)+'&'
	query = 'http://'+os.environ.get('SOLOIST_HOST')+'/run?'
	query += cmd[:-1]
	app.logger.warning('sendToSoloist - '+query)
	try:
		r = requests.get(query)
	except requests.exceptions.ConnectionError:
		app.logger.exception('sendToSoloist - Connection Error - '+query)
	return

###################################################
def sendToRadcor(activity,action):
# Sending command to radcor
	cmd = ''
	for key,val in activity.items():
		if val is not None:
			cmd += key+'='+str(val)+'&'
	query = 'http://'+os.environ.get('RADCOR_HOST')+'/{}?'.format(action)
	query += cmd[:-1]
	app.logger.warning('sendToRadcor - '+query)
	try:
		r = requests.get(query)
	except requests.exceptions.ConnectionError:
		app.logger.exception('sendToRadcor - Connection Error - '+query)
	return r.json()


################################
def whereDoIGo(activity,scene,section):

# Define the location in repository structure where file will be stored
# download_dir/dataset/yyyy_mm/pathrow
	datacube = activity['datacube']
	tileid = activity['tileid']
	start = activity['start']
	end = activity['end']
	dataset = scene['dataset']
	date = scene['date'].strftime('%Y-%m')
	pathrow = scene['pathrow']
	if section == 'Archive':
		dir = '/Repository/{}/{}'.format(section,dataset)
		dir += '/'+date
		dir += '/'+pathrow
	elif section == 'Warped':
		dir = '/Repository/{}/{}'.format(section,datacube)
		dir += '/'+tileid
		dir += '/{}-{}'.format(start,end)
	if not os.path.exists(dir):
		os.makedirs(dir)
	app.logger.warning('whereDoIGo - '+dir)
	return dir


###################################################
def checkWaitingDownload(activity):
	app.logger.warning('checkWaitingDownload - activity {}'.format(activity))
	sql = "UPDATE activities SET status = 'NOTDONE 'WHERE tsceneid = '{}' AND status = 'WAITING'".format(activity['tsceneid'])
	app.logger.warning('checkWaitingDownload - {}'.format(sql))
	do_command(sql)


###################################################
def downloadIfNecessary(activity):
# Get which scenes are related to this search
	params = "type = 'SCENE'"
	for key in ['datacube','tileid','start','end']:
		params += " AND {} = '{}'".format(key,activity[key])
			
	sql = "SELECT * FROM scenes WHERE {} ".format(params)
	scenes = do_query(sql)

# Create the new activity
	newactivity = {}
	newactivity['datacube'] = activity['datacube']
	newactivity['tileid'] = activity['tileid']
	newactivity['start'] = activity['start']
	newactivity['end'] = activity['end']
	newactivity['ttable'] = 'scenes'
	sceneDownloading = {}

	for scene in scenes:
		warped = whereDoIGo(activity,scene,'Warped')+'/'+scene['sceneid']+'_'+scene['band']+'.tif'
# If this scene is not a hdf file, lets warp it
		newactivity['tid'] = scene['id']
		newactivity['tsceneid'] = scene['sceneid']
		newactivity['band'] = scene['band']
		app.logger.warning('downloadIfNecessary - scene {}'.format(scene['sceneid']))
		if scene['dataset'] not in ['MOD13Q1','MYD13Q1']:
			newactivity['app'] = 'warp'
			newactivity['priority'] = 6
			newactivity['status'] = 'NOTDONE'
			file = '/vsicurl/'+ scene['link']
			file = scene['link'].replace('https://s3.amazonaws.com/datastorm-archive','/Repository/Archive')
# This scene is a hdf file, lets see what we have to do
		else:
# If this scene already exists on Archive, lets warp it
			file = whereDoIGo(activity,scene,'Archive')+'/'+scene['sceneid']+'_'+scene['band']+'.tif'
			newactivity['app'] = 'download'
			newactivity['priority'] = 2
# Check if this scene is on activities table being downloaded or waiting for download
			sql = "SELECT * FROM activities WHERE app = 'download' AND tsceneid = '{}' AND (status = 'DOING' OR status = 'NOTDONE' OR status = 'WAITING')".format(scene['sceneid'])
			waitingscenes = do_query(sql)
			if len(waitingscenes) > 0:
				sceneDownloading[scene['sceneid']] = True
# If this scene is not in sceneDownloading, download it now
			if scene['sceneid'] not in sceneDownloading:
				newactivity['status'] = 'NOTDONE'
				sceneDownloading[scene['sceneid']] = True
# If this scene is in sceneDownloading, leave it waiting for the end of the download
			else:
				newactivity['status'] = 'WAITING'
			do_upsert('activities',newactivity,['id','status','pstart','pend','elapsed','retcode','message'])
			newactivity['app'] = 'warp'
			newactivity['priority'] = 6
		sql = "UPDATE scenes SET file = '{}', warped = '{}' WHERE id = {}".format(file,warped,scene['id'])
		app.logger.warning('downloadIfNecessary - {} sceneDownloading {}'.format(sql,sceneDownloading))
		do_command(sql)
		app.logger.warning('downloadIfNecessary - newactivity {}'.format(newactivity))
		do_upsert('activities',newactivity,['id','status','pstart','pend','elapsed','retcode','message'])

	
###################################################
def doBlend(activity):

# Get the bands
	sql = "SELECT * FROM datacubes WHERE datacube = '{}'".format(activity['datacube'])
	result = do_query(sql)
	bands = result[0]['bands'].split(',')

# For all bands create the new activity
	for band in bands:
		if band == 'quality': continue
		newactivity = {}
		newactivity['app'] = 'blend'
		newactivity['datacube'] = activity['datacube']
		newactivity['tileid'] = activity['tileid']
		newactivity['start'] = activity['start']
		newactivity['end'] = activity['end']
		newactivity['band'] = band
		newactivity['priority'] = 4
		newactivity['status'] = 'NOTDONE'
		do_upsert('activities',newactivity,['id','status','pstart','pend','elapsed','retcode','message'])
	
###################################################
def shouldIMerge(activity):
# Check if all warpings for all bands related to this tile/period have finished

	params = "app = 'warp' AND status != 'DONE'"
	for key in ['datacube','tileid','start','end']:
		params += " AND {} = '{}'".format(key,activity[key])
	
	sql = "SELECT * FROM activities WHERE {} ".format(params)
	activities = do_query(sql)
# No warping still running, lets merge the images acquired in the same day
	app.logger.warning('shouldIMerge : {} activities still running'.format(len(activities)))
	if len(activities) == 0:
# Create the new activity
		newactivity = {}
		newactivity['app'] = 'merge'
		newactivity['datacube'] = activity['datacube']
		newactivity['tileid'] = activity['tileid']
		newactivity['start'] = activity['start']
		newactivity['end'] = activity['end']
		newactivity['priority'] = 5
		newactivity['status'] = 'NOTDONE'
		do_upsert('activities',newactivity,['id','status','pstart','pend','elapsed','retcode','message'])


###################################################
def shouldIPublish(activity):
# Check if all blendings related to this tile/period have finished

	params = "app = 'blend' AND status != 'DONE'"
	for key in ['datacube','tileid','start','end']:
		params += " AND {} = '{}'".format(key,activity[key])
	
	app.logger.warning('shouldIPublish : activity {}'.format(activity))
	sql = "SELECT * FROM activities WHERE {} ".format(params)
	activities = do_query(sql)
# No blending still running, lets publish this scene
	if len(activities) == 0:
# Create the new activity
		newactivity = {}
		newactivity['app'] = 'publish'
		newactivity['datacube'] = activity['datacube']
		newactivity['tileid'] = activity['tileid']
		newactivity['start'] = activity['start']
		newactivity['end'] = activity['end']
		newactivity['priority'] = 3
		newactivity['status'] = 'NOTDONE'
		do_upsert('activities',newactivity,['id','status','pstart','pend','elapsed','retcode','message'])


###################################################
def manage(activity):
	global MAX_THREADS,CUR_THREADS,ACTIVITIES
	app.logger.warning('manage start - lock : {} CUR_THREADS : {} ACTIVITIES : {} activity {}'.format(redis.get('lock'),CUR_THREADS,ACTIVITIES,activity))

# Create the critical region while database is modified. Leave it if sleeping time is greater than 10 units to avoid sleeping forever in a buggy situation
	countsleep = 0
	while (redis.get('lock') == b'1'):
		app.logger.warning('manage - sleep : {} activity {}'.format(countsleep,activity))
		time.sleep(0.5)
		countsleep += 1
		if countsleep > 10:
			redis.set('lock',0)
	redis.set('lock',1)

# Check if activity just finished or the flow is starting (id = -1)
	if int(activity['id']) >= 0:

		app.logger.warning('manage going to do_upsert lock : {} - id {}'.format(redis.get('lock'),activity['id']))
		do_upsert('activities',activity,['id','status','pstart','pend','elapsed','retcode','message'])
		CUR_THREADS -= 1
		if activity['app'] in ACTIVITIES:
			ACTIVITIES[activity['app']]['current'] -= 1
			
# activity just finished, lets see what must be done
	if activity['status'] == 'DONE':
# Create the next activities
# If search finished, download files if necessary (MODIS hdf files)
		if activity['app'] == 'search':
			downloadIfNecessary(activity)

# If download finished, check if other activities depend on this downloaded hdf file
		if activity['app'] == 'download':
			checkWaitingDownload(activity)

# If warping finished, check if all warp activities for this band are done
		if activity['app'] == 'warp':
			shouldIMerge(activity)

# If merge finished, blend scenes band by band
		if activity['app'] == 'merge':
			doBlend(activity)

# If all blending finished, publish scene
		if activity['app'] == 'blend':
			shouldIPublish(activity)

	elif activity['status'] == 'ERROR':
		app.logger.warning('manage - ERROR : activity {} ACTIVITIES {}'.format(activity,ACTIVITIES))

	sql = "SELECT * FROM activities WHERE status = 'NOTDONE' ORDER BY priority,id"
	result = do_query(sql)

	for newactivity in result:
		#app.logger.warning('manage loop newactivity - lock : {} CUR_THREADS : {} MAX_THREADS : {} activity {}'.format(redis.get('lock'),CUR_THREADS,MAX_THREADS,newactivity))

		if CUR_THREADS >= MAX_THREADS: break
		if newactivity['app'] in ACTIVITIES:
			if ACTIVITIES[newactivity['app']]['current'] >= ACTIVITIES[newactivity['app']]['maximum']:
				#app.logger.warning('manage - not yet activity {} {}'.format(ACTIVITIES[newactivity['app']],newactivity))
				continue
			else:
				ACTIVITIES[newactivity['app']]['current'] += 1
		newactivity['status'] = 'DOING'
		newactivity['elapsed'] = None
		step_start = time.time()
		newactivity['pstart'] = str(time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(step_start)))
		do_upsert('activities',newactivity,['id','status','pstart','pend','elapsed','retcode','message'])
# Avoid doing twice the download of the same hdf file
		#if newactivity['app'] in ACTIVITIES:
		#	noOtherLikeMe(newactivity)

		t = threading.Thread(target=sendToSoloist, args=(newactivity,))
		t.start()
		CUR_THREADS += 1
		#app.logger.warning('manage loop - lock : {} CUR_THREADS : {} {} MAX_THREADS {} {}'.format(redis.get('lock'),CUR_THREADS,type(CUR_THREADS),MAX_THREADS,type(MAX_THREADS)))

# Leave the critical region
	redis.set('lock',0)
	app.logger.warning('manage end - lock : {} CUR_THREADS : {} activity {}'.format(redis.get('lock'),CUR_THREADS,activity))
	return

###################################################
# @app.route('/process', methods=['GET'])
# def process():
# 	global MAX_THREADS
# 	msg = 'process - Maestro Processing:\n'
# 	name = request.args.get('datacube', 'LC8')
# 	msg += 'process - datacube is: {}\n'.format(name)
# 	pathrow = request.args.get('pr', None)

# # Manage lock
# 	lock = getLock()
# 	app.logger.warning('process - lock is: {}'.format(lock))
# 	msg += 'process - lock is: {}\n'.format(lock)

# # Manage MAX_THREADS
# 	nt = request.args.get('nt', None)
# 	msg += 'process - nt is: {}\n'.format(nt)
# 	app.logger.warning('process - nt is: {}'.format(nt))
# 	if MAX_THREADS is None:
# 		MAX_THREADS = int(os.environ.get('MAX_THREADS'))
# 		app.logger.warning('process - MAX_THREADS was None, now is: {} nt is: {}'.format(MAX_THREADS,nt))
# 	if nt is not None:
# 		MAX_THREADS = int(nt)
# 	app.logger.warning('process - MAX_THREADS is: {} nt is: {}'.format(MAX_THREADS,nt))
# 	msg += 'process - MAX_THREADS is: {} nt is: {}\n'.format(MAX_THREADS,nt)

# # Retrieve datacube info
# 	datacube = {}
# 	sql = "SELECT * FROM datacubes WHERE datacube = '{}'".format(name)
# 	result = do_query(sql)
# 	if len(result) > 0:
# 		for key,val in result[0].items():
# 			datacube[key] = val
# 	else:
# 		app.logger.warning('process - datacube {} not yet created'.format(name))
# 		return 'process - datacube {} not yet created'.format(name)

# # Get the time line for the datacube
# 	periodlist = decodePeriods(datacube['tschema'],datacube['start'],datacube['end'],datacube['step'])
# 	app.logger.warning('process - periodlist {}'.format(periodlist))

# # Get the requested period to be processed
# 	dstart = request.args.get('start', None)
# 	dend   = request.args.get('end', None)

# # Decode the requested tiles to be processed
# 	tileidlist = []
# 	if pathrow is not None:
# 		tileidlist = decodePathRow(pathrow)
# 	else:
# 		sql = "SELECT DISTINCT tileid FROM mosaics WHERE datacube = '{}'".format(datacube['datacube'])
# 		results = do_query(sql)
# 		for result in results:
# 			tileidlist.append(result['tileid'])
# 	app.logger.warning('process - tileidlist {}'.format(tileidlist))

# # Retrieve the mosaics to be processed
# 	for tileid in tileidlist:
# # Get the wrs information for this tile
# 		sql = "SELECT * FROM wrs WHERE name = '{}' AND tileid = '{}'".format(datacube['wrs'],tileid)
# 		wrs = do_query(sql,True)
# 		if wrs is None:
# 			continue

# 		for datekey in sorted(periodlist):
# 			requestedperiod = periodlist[datekey]
# 			for periodkey in requestedperiod:
# 				(basedate,startdate,enddate) = periodkey.split('_')
# 				if dstart is not None and startdate < dstart : continue
# 				if dend is not None and enddate > dend : continue

# # If this mosaic is not registered, insert it on mosaics table
# 				where = "WHERE datacube = '{}'".format(datacube['datacube'])
# 				where += " AND start >= '{}'".format(startdate)
# 				where += " AND end <= '{}'".format(enddate)
# 				where += " AND tileid = '{}'".format(tileid)
# 				sql = "SELECT * FROM mosaics {}".format(where)
# 				results = do_query(sql)
# 				if len(results) == 0:
# 					mosaics = {}
# 					mosaics['datacube'] = datacube['datacube']
# 					mosaics['tileid'] = tileid
# 					mosaics['start'] = startdate
# 					mosaics['end'] = enddate
# 					mosaics['numcol'] = int(round((wrs['xmax']-wrs['xmin'])/datacube['resx'],0))
# 					mosaics['numlin'] = int(round((wrs['ymax']-wrs['ymin'])/datacube['resy'],0))
# 					do_insert('mosaics',mosaics)


# # Process the mosaic
# 				results = do_query(sql)
# 				activity = {}
# 				activity['app'] = 'search'
# 				activity['status'] = 'NOTDONE'
# 				activity['priority'] = 1
# 				activity['ttable'] = 'mosaics'
# 				for result in results:
# 					activity['tid'] = result['id']
# 				for key in ['datacube','tileid','start','end']:
# 					activity[key] = result[key]
# 				app.logger.warning('process - activity {}'.format(activity))
# 				do_upsert('activities',activity,['id','status','pstart','pend','elapsed','retcode','message'])
# 				#return jsonify(activity)

# 	return start()

###################################################
@app.route('/process', methods=['GET'])
def process():
	global MAX_THREADS

	name = request.args.get('datacube', None)
	if not name: 
		return 'datacube name is required'

	pathrow = request.args.get('pr', None)
	if not pathrow:
		return 'pathrow is required'
	
	msg = 'process - Maestro ReProcessing:\n'
	msg += 'process - datacube is: {}\n'.format(name)

# Manage lock
	lock = getLock()
	app.logger.warning('process - lock is: {}'.format(lock))
	msg += 'process - lock is: {}\n'.format(lock)

# Manage MAX_THREADS
	nt = request.args.get('nt', None)
	msg += 'process - nt is: {}\n'.format(nt)
	app.logger.warning('process - nt is: {}'.format(nt))
	if MAX_THREADS is None:
		MAX_THREADS = int(os.environ.get('MAX_THREADS'))
		app.logger.warning('process - MAX_THREADS was None, now is: {} nt is: {}'.format(MAX_THREADS,nt))
	if nt is not None:
		MAX_THREADS = int(nt)
	app.logger.warning('process - MAX_THREADS is: {} nt is: {}'.format(MAX_THREADS,nt))
	msg += 'process - MAX_THREADS is: {} nt is: {}\n'.format(MAX_THREADS,nt)

# Retrieve datacube info
	datacube = {}
	sql = "SELECT * FROM datacubes WHERE datacube = '{}'".format(name)
	result = do_query(sql)
	if len(result) > 0:
		for key,val in result[0].items():
			datacube[key] = val
	else:
		app.logger.warning('process - datacube {} not yet created'.format(name))
		return 'process - datacube {} not yet created'.format(name)

# define new range temporal of the cube
	dstart = datetime.datetime.strptime(request.args.get('start'), '%Y-%m-%d').date() if request.args.get('start') else None
	dend = datetime.datetime.strptime(request.args.get('end'), '%Y-%m-%d').date() if request.args.get('end') else None
	if not dstart or not dend:
		return 'date start and date end is required'

# Get the time line for the datacube
	end_period = datacube['end'] if datacube['end'] >= dend else dend
	periodlist = decodePeriods(datacube['tschema'],datacube['start'],end_period,datacube['step'])
	app.logger.warning('process - periodlist {}'.format(periodlist))

# Decode the requested tiles to be processed
	tileidlist = []
	if pathrow is not None:
		tileidlist = decodePathRow(pathrow)
	else:
		sql = "SELECT DISTINCT tileid FROM mosaics WHERE datacube = '{}'".format(datacube['datacube'])
		results = do_query(sql)
		for result in results:
			tileidlist.append(result['tileid'])
	app.logger.warning('process - tileidlist {}'.format(tileidlist))

# Retrieve the mosaics to be processed
	for tileid in tileidlist:
# Get the wrs information for this tile
		sql = "SELECT * FROM wrs WHERE name = '{}' AND tileid = '{}'".format(datacube['wrs'],tileid)
		wrs = do_query(sql,True)
		if wrs is None:
			continue

		for datekey in sorted(periodlist):
			requestedperiod = periodlist[datekey]
			for periodkey in requestedperiod:
				(basedate,startdate,enddate) = periodkey.split('_')
				if datetime.datetime.strptime(startdate, '%Y-%m-%d').date() < dstart or datetime.datetime.strptime(enddate, '%Y-%m-%d').date() > dend:
					print( "jumped", datetime.datetime.strptime(startdate, '%Y-%m-%d').date() )
					continue

# If this mosaic is not registered, insert it on mosaics table
				where = "WHERE datacube = '{}'".format(datacube['datacube'])
				where += " AND start >= '{}'".format(startdate)
				where += " AND end <= '{}'".format(enddate)
				where += " AND tileid = '{}'".format(tileid)
				sql = "SELECT * FROM mosaics {}".format(where)
				results = do_query(sql)
				if len(results) == 0:
					mosaics = {}
					mosaics['datacube'] = datacube['datacube']
					mosaics['tileid'] = tileid
					mosaics['start'] = startdate
					mosaics['end'] = enddate
					mosaics['numcol'] = int(round((wrs['xmax']-wrs['xmin'])/datacube['resx'],0))
					mosaics['numlin'] = int(round((wrs['ymax']-wrs['ymin'])/datacube['resy'],0))
					do_insert('mosaics',mosaics)

# Process the mosaic
				results = do_query(sql)
				activity = {}
				activity['app'] = 'search'
				activity['status'] = 'NOTDONE'
				activity['priority'] = 1
				activity['ttable'] = 'mosaics'
				for result in results:
					activity['tid'] = result['id']
				for key in ['datacube','tileid','start','end']:
					activity[key] = result[key]

# remove activity if exists
				sql_delete = "DELETE FROM activities WHERE datacube = '{}' AND tileid = '{}' and start = '{}' and end = '{}'".format(
					activity['datacube'], activity['tileid'], activity['start'], activity['end'])
				do_command(sql_delete)

# insert Activity with search
				app.logger.warning('process - activity {}'.format(activity))
				do_upsert('activities',activity,['id','status','pstart','pend','elapsed','retcode','message'])
				#return jsonify(activity)
	if dend > datacube['end']:
		sql_update = "UPDATE datacubes SET end = '{}' WHERE datacube = '{}'".format(dend, activity['datacube'])
		do_command(sql_update)

	return start()

###################################################
@app.route('/radcor', methods=['GET'])
def radcor():
	name = request.args.get('datacube', 'LC8')
	action = request.args.get('action', 'search')
	limit = request.args.get('limit', 200)
	cloud = request.args.get('cloud', 10)
	pr = request.args.get('pr', None)

# Retrieve datacube info
	datacube = {}
	sql = "SELECT * FROM datacubes WHERE datacube = '{}'".format(name)
	result = do_query(sql)
	if len(result) > 0:
		for key,val in result[0].items():
			datacube[key] = val
	else:
		app.logger.warning('radcor - datacube {} not yet created'.format(name))
		return 'radcor - datacube {} not yet created'.format(name)
	
	activity = {}
	sats = []
	satsen = datacube['satsen'].split(',')
	for sat in satsen:
		if sat == 'S2SR':
			sats.append('S2')
		if sat == 'LC8SR':
			sats.append('LC8')
	sats = ",".join(sats)
	
	activity['sat'] = sats
	activity['limit'] = limit
	activity['cloud'] = cloud

# Get the time line for the datacube
	periodlist = decodePeriods(datacube['tschema'],datacube['start'],datacube['end'],datacube['step'])
# Get the requested period to be processed
	dstart = request.args.get('start', None)
	dend   = request.args.get('end', None)

# Decode the requested tiles to be processed
	tileidlist = []
	if pr is not None:
		pathrowlist = decodePathRow(pr)
		for tileid in pathrowlist:
			tileidlist.append(tileid)
			app.logger.warning('process - tileid {}'.format(tileid))
	else:
		sql = "SELECT DISTINCT tileid FROM mosaics WHERE datacube = '{}'".format(name)
		results = do_query(sql)
		for result in results:
			tileidlist.append(result['tileid'])

# Retrieve the mosaics to be processed
	scenes = {}
	total = {}
	for tileid in tileidlist:
		sql = "SELECT * FROM wrs WHERE tileid = '{}'".format(tileid)
		results = do_query(sql)
		if len(results) == 0:
			continue
		activity['w'] = results[0]['lonmin']
		activity['e'] = results[0]['lonmax']
		activity['s'] = results[0]['latmin']
		activity['n'] = results[0]['latmax']
		for datekey in sorted(periodlist):
			requestedperiod = periodlist[datekey]
			for periodkey in requestedperiod:
				(basedate,startdate,enddate) = periodkey.split('_')
				app.logger.warning('radcor - startdate {} dstart {} '.format(startdate,dstart))
				if dstart is not None and startdate < dstart : continue
				if dend is not None and enddate > dend : continue
				activity['start'] = startdate
				activity['end'] = enddate
				app.logger.warning('radcor - activity {} '.format(activity))
				scene = sendToRadcor(activity,action)
				if action == 'search':
					scene = filter(scene)
					scenes['{}-{}-{}'.format(tileid,startdate,enddate)] = scene
				else:
					scenes.update(scene)
				total['{}-{}-{}'.format(tileid,startdate,enddate)]= len(scene)
	if action == 'search':
		scenes['Results'] = total
	return jsonify(scenes)

###################################################
def filter(scenes,tags=['cloud','date']):
	newscenes = {}
	for sceneid in scenes:
		scene = scenes[sceneid]
		newscenes[sceneid] = {}
		for tag in tags:
			newscenes[sceneid][tag] = scene[tag]
	return newscenes

###################################################
@app.route('/blend', methods=['GET'])
def blend():
	global MAX_THREADS,CUR_THREADS
	msg = 'blend - Maestro Processing:\n'
	name = request.args.get('datacube', 'LC8')
	msg += 'blend - datacube is: {}\n'.format(name)
	pr = request.args.get('pr', None)
	if pr is None:
		return 'Enter pr\n'

# Manage MAX_THREADS
	CUR_THREADS = 0
	nt = request.args.get('nt', None)
	msg += 'blend - nt is: {}\n'.format(nt)
	app.logger.warning('blend - nt is: {}'.format(nt))
	if MAX_THREADS is None:
		MAX_THREADS = int(os.environ.get('MAX_THREADS'))
		app.logger.warning('blend - MAX_THREADS was None, now is: {} nt is: {}'.format(MAX_THREADS,nt))
	if nt is not None:
		MAX_THREADS = int(nt)
	app.logger.warning('blend - MAX_THREADS is: {} nt is: {}'.format(MAX_THREADS,nt))
	msg += 'blend - MAX_THREADS is: {} nt is: {}\n'.format(MAX_THREADS,nt)

# Retrieve datacube info
	datacube = {}
	sql = "SELECT * FROM datacubes WHERE datacube = '{}'".format(name)
	result = do_query(sql)
	if len(result) > 0:
		for key,val in result[0].items():
			datacube[key] = val
	else:
		app.logger.warning('blend - datacube {} not yet created'.format(name))
		return 'blend - datacube {} not yet created'.format(name)
	app.logger.warning('blend - datacube {} '.format(datacube))

# Get the requested period to be processed
	dstart = request.args.get('start', None)
	dend   = request.args.get('end', None)

# Decode the requested tiles to be processed
	pathrowlist = decodePathRow(pr)

# Get the time line for the datacube
	periodlist = decodePeriods(datacube['tschema'],datacube['start'],datacube['end'],datacube['step'])
	for datekey in sorted(periodlist):
		requestedperiod = periodlist[datekey]
		for periodkey in requestedperiod:
			(basedate,startdate,enddate) = periodkey.split('_')
			if dstart is not None and enddate < dstart:
				continue
			if dend is not None and startdate > dend:
				continue
			app.logger.warning('blend - datekey {} period {}:{}->{}'.format(datekey,basedate,startdate,enddate))

# Process the requested tiles
			for pathrow in pathrowlist:
				activity = {}
				activity['datacube'] = datacube['datacube']
				activity['tileid'] = pathrow
				activity['start'] = startdate
				activity['end'] = enddate
				doBlend(activity)

	return start()


###################################################
@app.route('/publish', methods=['GET'])
def publish():
	global MAX_THREADS,CUR_THREADS
	msg = 'blend - Maestro Processing:\n'
	name = request.args.get('datacube', 'LC8')
	msg += 'publish - datacube is: {}\n'.format(name)
	pr = request.args.get('pr', None)
	if pr is None:
		return 'Enter pr\n'

# Manage MAX_THREADS
	CUR_THREADS = 0
	nt = request.args.get('nt', None)
	msg += 'publish - nt is: {}\n'.format(nt)
	app.logger.warning('blend - nt is: {}'.format(nt))
	if MAX_THREADS is None:
		MAX_THREADS = int(os.environ.get('MAX_THREADS'))
		app.logger.warning('blend - MAX_THREADS was None, now is: {} nt is: {}'.format(MAX_THREADS,nt))
	if nt is not None:
		MAX_THREADS = int(nt)
	app.logger.warning('publish - MAX_THREADS is: {} nt is: {}'.format(MAX_THREADS,nt))
	msg += 'publish - MAX_THREADS is: {} nt is: {}\n'.format(MAX_THREADS,nt)

# Retrieve datacube info
	datacube = {}
	sql = "SELECT * FROM datacubes WHERE datacube = '{}'".format(name)
	result = do_query(sql)
	if len(result) > 0:
		for key,val in result[0].items():
			datacube[key] = val
	else:
		app.logger.warning('publish - datacube {} not yet created'.format(name))
		return 'blend - datacube {} not yet created'.format(name)
	app.logger.warning('publish - datacube {} '.format(datacube))

# Get the requested period to be processed
	dstart = request.args.get('start', None)
	dend   = request.args.get('end', None)

# Decode the requested tiles to be processed
	pathrowlist = decodePathRow(pr)

# Get the time line for the datacube
	periodlist = decodePeriods(datacube['tschema'],datacube['start'],datacube['end'],datacube['step'])
	for datekey in sorted(periodlist):
		requestedperiod = periodlist[datekey]
		for periodkey in requestedperiod:
			(basedate,startdate,enddate) = periodkey.split('_')
			if dstart is not None and enddate < dstart:
				continue
			if dend is not None and startdate > dend:
				continue
			app.logger.warning('blend - datekey {} period {}:{}->{}'.format(datekey,basedate,startdate,enddate))

# Process the requested tiles
			for pathrow in pathrowlist:
				activity = {}
				activity['datacube'] = datacube['datacube']
				activity['tileid'] = pathrow
				activity['start'] = startdate
				activity['end'] = enddate
				shouldIPublish(activity)

	return start()


###################################################
@app.route('/manage', methods=['GET'])
def decodeRequest():
	activity = {}
	activity['id'] = request.args.get('id', None)
	activity['app'] = request.args.get('app', None)
	activity['priority'] = request.args.get('priority', None)
	activity['datacube'] = request.args.get('datacube', None)
	activity['tileid'] = request.args.get('tileid', None)
	activity['start'] = request.args.get('start', None)
	activity['end'] = request.args.get('end', None)
	activity['ttable'] = request.args.get('ttable', None)
	activity['tid'] = request.args.get('tid', None)
	activity['tsceneid'] = request.args.get('tsceneid', None)
	activity['band'] = request.args.get('band', None)
	activity['status'] = request.args.get('status', None)
	activity['pstart'] = request.args.get('pstart', None)
	activity['pend'] = request.args.get('pend', None)
	activity['elapsed'] = request.args.get('elapsed', None)
	activity['retcode'] = request.args.get('retcode', None)
	activity['message'] = request.args.get('message', None)
	manage(activity)
	return 'OK\n'


###################################################
@app.route('/clean', methods=['GET'])
def clean():
	global MAX_THREADS,CUR_THREADS
	msg = 'Maestro Processing:\n'
	redis.set('lock',0)
	lock = redis.get('lock')
	msg += 'lock is: {}\n'.format(lock)
	msg += 'MAX_THREADS is: {}\n'.format(MAX_THREADS)
	CUR_THREADS = 0
	msg += 'CUR_THREADS is: {}\n'.format(CUR_THREADS)
	for key in ACTIVITIES:
		ACTIVITIES[key]['current'] = 0
	msg += 'ACTIVITIES is: {}\n'.format(ACTIVITIES)
	
	name = request.args.get('datacube', None)

	if name is None:
		sql = "TRUNCATE mosaics"
	else:
		sql = "DELETE FROM mosaics WHERE datacube = '{}'".format(name)
	do_command(sql)
	if name is None:
		sql = "TRUNCATE scenes"
	else:
		sql = "DELETE FROM scenes WHERE datacube = '{}'".format(name)
	do_command(sql)
	if name is None:
		sql = "TRUNCATE products"
	else:
		sql = "DELETE FROM products WHERE datacube = '{}'".format(name)
	do_command(sql)
	if name is None:
		sql = "TRUNCATE activities"
	else:
		sql = "DELETE FROM activities WHERE datacube = '{}'".format(name)
	do_command(sql)
	return msg


###################################################
@app.route('/stats', methods=['GET'])
def stats():
	sql = "SELECT * FROM activities WHERE status = 'DONE'"
	result = do_query(sql)
	appstats = {}
	for activity in result:
		elapsed = int(activity['elapsed'].total_seconds())
		if activity['app'] not in appstats:
			appstats[activity['app']] = []
		appstats[activity['app']].append(elapsed)
	app.logger.warning('stats - {}'.format(appstats))
	wb = openpyxl.Workbook()
	sheet1 = wb.active
	sheet1.title = 'Performance'
	fieldnames = []
	fieldnames.append('Activity')
	fieldnames.append('Min')
	fieldnames.append('Max')
	fieldnames.append('Count')
	fieldnames.append('Mean')
	ci = 1
	for field in fieldnames:
		sheet1.cell(row=1, column=ci).value = field
		ci += 1
	ri = 2
	for key in appstats:
		appstat = appstats[key]
		tmin = min(appstat)
		tmax = max(appstat)
		tcount = len(appstat)
		tmean = sum(appstat)/tcount
		app.logger.warning('{} - {} - {} - {} - {}'.format(key,tmin,tmax,tcount,tmean))
		sheet1.cell(row=ri, column=1).value = key
		sheet1.cell(row=ri, column=2).value = tmin
		sheet1.cell(row=ri, column=3).value = tmax
		sheet1.cell(row=ri, column=4).value = tcount
		sheet1.cell(row=ri, column=5).value = tmean
		ri += 1

	datacube = request.args.get('datacube', 'prodes')
	tile = request.args.get('tile', '226064')

	sql = "SELECT DISTINCT sceneid FROM scenes WHERE type = 'SCENE' AND datacube = '{}' AND tileid = '{}'".format(datacube,tile)
	sql = "SELECT * FROM scenes"
	result = do_query(sql)
	scenes = {}
	for scene in result:
		datacube = scene['datacube']
		tileid = scene['tileid']
		type = scene['type']
		dataset = scene['dataset']
		start = '{}'.format(scene['start'])
		band = scene['band']
		file = scene['warped']
		if not os.path.exists(file): continue
		size = int(round(os.path.getsize(file)/1024))
		if datacube not in scenes:
			scenes[datacube] = {}
		if tileid not in scenes[datacube]:
			scenes[datacube][tileid] = {}
		if start not in scenes[datacube][tileid]:
			scenes[datacube][tileid][start] = {}
		if type not in scenes[datacube][tileid][start]:
			scenes[datacube][tileid][start][type] = {}
		if band not in scenes[datacube][tileid][start][type]:
			scenes[datacube][tileid][start][type][band] = {}
		if dataset not in scenes[datacube][tileid][start][type][band]:
			scenes[datacube][tileid][start][type][band][dataset] = {}
			scenes[datacube][tileid][start][type][band][dataset]['count'] = 0
			scenes[datacube][tileid][start][type][band][dataset]['size'] = 0
			scenes[datacube][tileid][start][type][band][dataset]['files'] = []
		bfile = os.path.basename(file)
		if bfile in scenes[datacube][tileid][start][type][band][dataset]['files']: continue
		scenes[datacube][tileid][start][type][band][dataset]['count'] += 1
		scenes[datacube][tileid][start][type][band][dataset]['size'] += size
		scenes[datacube][tileid][start][type][band][dataset]['files'].append(bfile)

	app.logger.warning('scenes - {}'.format(scenes))
	sheet2 = wb.create_sheet()
	sheet2.title = 'Storage'
	fieldnames = []
	fieldnames.append('Datacube')
	fieldnames.append('Tile')
	fieldnames.append('Start')
	fieldnames.append('Type')
	fieldnames.append('Band')
	fieldnames.append('Dataset')
	fieldnames.append('Count')
	fieldnames.append('Size (KB)')
	ci = 1
	for field in fieldnames:
		sheet2.cell(row=1, column=ci).value = field
		ci += 1
	ri = 2
	for datacube in scenes:
		for tileid in scenes[datacube]:
			for start in scenes[datacube][tileid]:
				for type in scenes[datacube][tileid][start]:
					for band in scenes[datacube][tileid][start][type]:
						for dataset in scenes[datacube][tileid][start][type][band]:
							sheet2.cell(row=ri, column=1).value = datacube
							sheet2.cell(row=ri, column=2).value = tileid
							sheet2.cell(row=ri, column=3).value = start
							sheet2.cell(row=ri, column=4).value = type
							sheet2.cell(row=ri, column=5).value = band
							sheet2.cell(row=ri, column=6).value = dataset
							sheet2.cell(row=ri, column=7).value = scenes[datacube][tileid][start][type][band][dataset]['count']
							sheet2.cell(row=ri, column=8).value = scenes[datacube][tileid][start][type][band][dataset]['size']
							ri += 1

# Create xls for this scene
	xlsfilename = 'stats.xlsx'
	wb.save(xlsfilename)

	return jsonify(scenes)
	
###################################################
@app.route('/inspect', methods=['GET'])
def inspect():
	global MAX_THREADS,CUR_THREADS
	msg = 'Maestro Processing:\n'
	status = request.args.get('status', None)
	cube_name = request.args.get('cubename', None)
	lock = getLock()
	msg += 'lock is: {}\n'.format(lock)
	msg += 'MAX_THREADS is: {}\n'.format(MAX_THREADS)
	msg += 'CUR_THREADS is: {}\n'.format(CUR_THREADS)
	msg += 'ACTIVITIES is: {}\n'.format(ACTIVITIES)
	
	sql = "SELECT * FROM activities "
	if status is not None or cube_name is not None:
		sql += "WHERE "

		if status is not None:
			sql += "status = '{}' ".format(status)
			if cube_name is not None:
				sql += "and datacube = '{}' ".format(cube_name)

		elif cube_name is not None:
			sql += "datacube = '{}' ".format(cube_name)

	sql += "ORDER BY id"
	result = do_query(sql)

	for activity in result:
		msg += '{} - {} - {} - {} {} {} -> {}\n'.format(activity['id'],activity['app'],activity['status'],activity['datacube'],activity['tileid'],activity['start'],activity['end'])
	return msg
	

###################################################
@app.route('/start', methods=['GET'])
def start():
	activity = {}
	activity['id'] = -1
	activity['app'] = 'START'
	activity['status'] = 'START'
	manage(activity)
	return 'OK\n'

###################################################
@app.route('/restart', methods=['GET'])
def restart():
	global MAX_THREADS,CUR_THREADS,ACTIVITIES
	msg = 'Maestro restarting:\n'
	id = request.args.get('id', None)
	if id is None:
		sql = "UPDATE activities SET status='NOTDONE' WHERE (status = 'ERROR' OR status = 'DOING')"
	else:
		sql = "UPDATE activities SET status='NOTDONE' WHERE id = {}".format(id)
	do_command(sql)
	msg += 'sql - {}\n'.format(sql)
	CUR_THREADS = 0
	ACTIVITIES = {'search':{'current':0,'maximum':8},'download':{'current':0,'maximum':8},'warp':{'current':0,'maximum':16},'merge':{'current':0,'maximum':16},'blend':{'current':0,'maximum':8}}
	msg += 'ACTIVITIES - {}\n'.format(ACTIVITIES)

	start()
	return msg

###################################################
@app.route('/pause', methods=['GET'])
def pause():
        global MAX_THREADS,CUR_THREADS,ACTIVITIES
        msg = 'Maestro restarting:\n'
        id = request.args.get('id', None)
        sql = "UPDATE activities SET status='SUSPEND' WHERE status = 'NOTDONE'"
        do_command(sql)
        msg += 'sql - {}\n'.format(sql)
        msg += 'ACTIVITIES - {}\n'.format(ACTIVITIES)

        return msg

####################################################
@app.route('/do2ch', methods=['GET'])
#Update status from DOING (zombies) to CHECK
def do2ch():
	global MAX_THREADS,CUR_THREADS,ACTIVITIES
	msg = 'Maestro restarting:\n Change stucked DOING processes to CHECK status\n'
	sql = "UPDATE activities SET status='CHECK' WHERE status = 'DOING'"
	
	do_command(sql)
	msg += 'sql - {}\n'.format(sql)
	CUR_THREADS = 0
	ACTIVITIES = {'search':{'current':0,'maximum':8},'download':{'current':0,'maximum':8},'warp':{'current':0,'maximum':16},'merge':{'current':0,'maximum':16},'blend':{'current':0,'maximum':8}}
	msg += 'ACTIVITIES - {}\n'.format(ACTIVITIES)

	start()
	return msg
	

###################################################
@app.route('/deldice', methods=['GET'])
def deldice():
	datacube = request.args.get('datacube', None)
	tileid = request.args.get('tileid', None)
	startdate = request.args.get('startdate', None)
	msg = 'Delete rows from a specific cube/tile/date\n'
	if datacube is None or tileid is None or startdate is None:
		msg = 'Supply all variables (datacube, tileid, startdate)\n'
	else:
		sql = "DELETE FROM scenes WHERE datacube = '{0}' AND tileid ='{1}' AND start = '{2}'".format(datacube,tileid,startdate)
		do_command(sql)
		sql = "DELETE FROM activities WHERE datacube = '{0}' AND tileid ='{1}' AND start = '{2}'".format(datacube,tileid,startdate)
		do_command(sql)
		sql = "DELETE FROM products WHERE datacube = '{0}' AND tileid ='{1}' AND start = '{2}'".format(datacube,tileid,startdate)
		do_command(sql)
		sql = "DELETE FROM qlook WHERE datacube = '{0}' AND tileid ='{1}' AND start = '{2}'".format(datacube,tileid,startdate)
		do_command(sql)

	# Delete all related files
		warped = '/Repository/Warped/{0}/{1}/{2}*'.format(datacube,tileid,startdate)
		msg += '{}\n'.format(warped)
		mosaic = '/Repository/Mosaic/{0}/{1}/{2}*'.format(datacube,tileid,startdate)
		msg += '{}\n'.format(mosaic)
		if os.path.exists(warped):
			shutil.rmtree(warped)
		if os.path.exists(mosaic):
			shutil.rmtree(mosaic)
		resp = jsonify({'code': 200, 'message': 'Normal execution'})
		resp.status_code = 200
		#return resp
	return msg

	start()
	return msg
	
###################################################
@app.errorhandler(400)
def handle_bad_request(e):
	resp = jsonify({'code': 400, 'message': 'Bad Request - {}'.format(e.description)})
	resp.status_code = 400
	resp.headers.add('Access-Control-Allow-Origin', '*')
	return resp


@app.errorhandler(404)
def handle_page_not_found(e):
	resp = jsonify({'code': 404, 'message': 'Page not found'})
	app.logger.info('code: 404 : Page not found')
	resp.status_code = 404
	resp.headers.add('Access-Control-Allow-Origin', '*')
	return resp


@app.errorhandler(500)
def handle_api_error(e):
	resp = jsonify({'code': 500, 'message': 'Internal Server Error'})
	resp.status_code = 500
	resp.headers.add('Access-Control-Allow-Origin', '*')
	return resp


@app.errorhandler(502)
def handle_bad_gateway_error(e):
	resp = jsonify({'code': 502, 'message': 'Bad Gateway'})
	resp.status_code = 502
	resp.headers.add('Access-Control-Allow-Origin', '*')
	return resp


@app.errorhandler(503)
def handle_service_unavailable_error(e):
	resp = jsonify({'code': 503, 'message': 'Service Unavailable'})
	resp.status_code = 503
	resp.headers.add('Access-Control-Allow-Origin', '*')
	return resp


@app.errorhandler(Exception)
def handle_exception(e):
	app.logger.exception(e)
	resp = jsonify({'code': 500, 'message': 'Internal Server Error'})
	resp.status_code = 500
	resp.headers.add('Access-Control-Allow-Origin', '*')
	return resp

